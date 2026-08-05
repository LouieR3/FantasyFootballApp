import os as _os, sys as _sys
_d = _os.path.dirname(_os.path.abspath(__file__))
while _d != _os.path.dirname(_d) and not _os.path.exists(_os.path.join(_d, 'paths.py')):
    _d = _os.path.dirname(_d)
_sys.path.insert(0, _d)
from credentials import CRED
from espn_api.football import League
import pandas as pd
import os
from openpyxl import load_workbook
from paths import DATA_DIR, LEAGUES_DIR

def winless_record_chances(leagues, years):
    """
    Analyzes the percentage chance of making the playoffs for teams starting winless at 0-1 through 0-7.

    Parameters:
    - leagues (list): A list of dictionaries containing league configurations (league_id, espn_s2, swid, name).
    - years (list): List of years to analyze.

    Returns:
    - pd.DataFrame: A DataFrame summarizing the playoff chances for each winless record.
    """
    def create_playoff_df(leagues, years):
        
        # Initialize an empty list to store all playoff data
        combined_playoff_dfs = []

        # Loop through each league
        for league_config in leagues:
            league_id = league_config['league_id']
            espn_s2 = league_config['espn_s2']
            swid = league_config['swid']
            league_name = league_config['name']

            for year in years:
                print(f"Processing league: {league_name}, year: {year}")

                # Instantiate the league object for the current year
                try:
                    league = League(league_id=league_id, year=year, espn_s2=espn_s2, swid=swid)
                except Exception as e:
                    # print(f"Error initializing league {league_name} for year {year}: {e}")
                    continue

                # Get the league name and construct the file path
                file_path = f"{LEAGUES_DIR}/{league_name} {year}.xlsx"

                # Check if the file exists
                if not os.path.exists(file_path):
                    # print(f"File not found for year {year}: {file_path}. Skipping this year.")
                    continue

                try:
                    # Load the workbook
                    workbook = load_workbook(file_path, read_only=True)

                    # Check if "Playoff Results" sheet exists
                    if "Playoff Results" not in workbook.sheetnames:
                        # print(f"Skipping {file_path}: 'Playoff Results' sheet not found.")
                        continue

                    # Read the Playoff Results sheet into a DataFrame
                    playoff_df = pd.read_excel(file_path, sheet_name="Playoff Results")
                    playoff_df['Year'] = year  # Add year for identification
                    playoff_df['League'] = league_name  # Add league name for identification
                    # basename only - an absolute path here would be machine-specific
                    # and would break consumers that match on this column
                    playoff_df['File Name'] = os.path.basename(file_path)
                    combined_playoff_dfs.append(playoff_df)

                    print(f"Processed {file_path}: 'Playoff Results' sheet loaded.")

                except Exception as e:
                    print(f"Error processing {file_path}: {e}")

        # Combine all playoff DataFrames into one
        if not combined_playoff_dfs:
            # print("No playoff data found.")
            return pd.DataFrame(), pd.DataFrame()  # Return empty DataFrames if no data is found

        all_playoff_dfs = pd.concat(combined_playoff_dfs, ignore_index=True)
        all_playoff_dfs.to_csv(f"{DATA_DIR}/all_playoff_dfs.csv", index=False)
        print(f"Combined playoff DataFrame:\n{all_playoff_dfs}")
        return all_playoff_dfs
    # all_playoff_dfs = create_playoff_df(leagues, years)
    all_playoff_dfs = pd.read_csv(f"{DATA_DIR}/all_playoff_dfs.csv")
    if all_playoff_dfs.empty:
        return pd.DataFrame(), pd.DataFrame()
    print(f"All playoff DataFrame:\n{all_playoff_dfs}")

    # Analyze playoff chances for teams starting 0-1 through 0-7
    playoff_chances = []
    for league_config in leagues:
        league_id = league_config['league_id']
        espn_s2 = league_config['espn_s2']
        swid = league_config['swid']
        league_name = league_config['name']

        for year in years:
            try:
                league = League(league_id=league_id, year=year, espn_s2=espn_s2, swid=swid)
            except Exception as e:
                print(f"Error initializing league {league_name} for year {year}: {e}")
                continue

            # Filter the playoff DataFrame for the current year and league
            year_playoff_df = all_playoff_dfs[
                (all_playoff_dfs['Year'] == year) & (all_playoff_dfs['League'] == league_name)
            ]

            # Combine Team 1 and Team 2 columns to get all teams that participated in the playoffs
            playoff_teams = pd.concat([year_playoff_df['Team 1'], year_playoff_df['Team 2']])
            playoff_teams = playoff_teams[playoff_teams != 'Bye'].unique()

            print(f"Playoff teams for league: {league_name}, year {year}: {playoff_teams}")

            # Loop through winless records from 0-1 to 0-7
            for winless_record in range(1, 8):
                # Get teams with the current winless record
                winless_teams = [
                    team.team_name
                    for team in league.teams
                    if team.outcomes[:winless_record].count('L') == winless_record
                ]
                print(f"Winless teams with record 0-{winless_record} for league: {league_name}, year {year}: {winless_teams}")

                # Stop searching if no teams match the current winless record
                if not winless_teams:
                    break

                # Calculate the percentage of winless teams that made the playoffs
                made_playoffs = [team for team in winless_teams if team in playoff_teams]
                playoff_percentage = len(made_playoffs) / len(winless_teams) * 100

                # Append the result to the playoff chances list
                playoff_chances.append({
                    "League": league_name,
                    "Year": year,
                    "Winless Record": f"0-{winless_record}",
                    "Total Teams": len(winless_teams),
                    "Made Playoffs": len(made_playoffs),
                    "Playoff Percentage": playoff_percentage
                })

    # Convert the playoff chances list to a DataFrame
    playoff_chances_df = pd.DataFrame(playoff_chances)

    # Aggregate totals, made playoffs, and playoff percentage by year and league
    aggregated_df = playoff_chances_df.groupby(['Winless Record']).agg({
        'Total Teams': 'sum',
        'Made Playoffs': 'sum'
    }).reset_index()

    # Calculate the aggregated playoff percentage
    aggregated_df['Playoff Percentage'] = (aggregated_df['Made Playoffs'] / aggregated_df['Total Teams']) * 100

    # Return the final DataFrames
    return aggregated_df, playoff_chances_df

def undefeated_record_chances(leagues, years):
    """
    Analyzes the percentage chance of making the playoffs for teams starting winless at 0-1 through 0-7.

    Parameters:
    - leagues (list): A list of dictionaries containing league configurations (league_id, espn_s2, swid, name).
    - years (list): List of years to analyze.

    Returns:
    - pd.DataFrame: A DataFrame summarizing the playoff chances for each winless record.
    """
    def create_playoff_df(leagues, years):
        
        # Initialize an empty list to store all playoff data
        combined_playoff_dfs = []

        # Loop through each league
        for league_config in leagues:
            league_id = league_config['league_id']
            espn_s2 = league_config['espn_s2']
            swid = league_config['swid']
            league_name = league_config['name']

            for year in years:
                print(f"Processing league: {league_name}, year: {year}")

                # Instantiate the league object for the current year
                try:
                    league = League(league_id=league_id, year=year, espn_s2=espn_s2, swid=swid)
                except Exception as e:
                    # print(f"Error initializing league {league_name} for year {year}: {e}")
                    continue

                # Get the league name and construct the file path
                file_path = f"{LEAGUES_DIR}/{league_name} {year}.xlsx"

                # Check if the file exists
                if not os.path.exists(file_path):
                    # print(f"File not found for year {year}: {file_path}. Skipping this year.")
                    continue

                try:
                    # Load the workbook
                    workbook = load_workbook(file_path, read_only=True)

                    # Check if "Playoff Results" sheet exists
                    if "Playoff Results" not in workbook.sheetnames:
                        # print(f"Skipping {file_path}: 'Playoff Results' sheet not found.")
                        continue

                    # Read the Playoff Results sheet into a DataFrame
                    playoff_df = pd.read_excel(file_path, sheet_name="Playoff Results")
                    playoff_df['Year'] = year  # Add year for identification
                    playoff_df['League'] = league_name  # Add league name for identification
                    # basename only - an absolute path here would be machine-specific
                    # and would break consumers that match on this column
                    playoff_df['File Name'] = os.path.basename(file_path)
                    combined_playoff_dfs.append(playoff_df)

                    print(f"Processed {file_path}: 'Playoff Results' sheet loaded.")

                except Exception as e:
                    print(f"Error processing {file_path}: {e}")

        # Combine all playoff DataFrames into one
        if not combined_playoff_dfs:
            # print("No playoff data found.")
            return pd.DataFrame(), pd.DataFrame()  # Return empty DataFrames if no data is found

        all_playoff_dfs = pd.concat(combined_playoff_dfs, ignore_index=True)
        all_playoff_dfs.to_csv(f"{DATA_DIR}/all_playoff_dfs.csv", index=False)
        print(f"Combined playoff DataFrame:\n{all_playoff_dfs}")
        return all_playoff_dfs
    # all_playoff_dfs = create_playoff_df(leagues, years)
    all_playoff_dfs = pd.read_csv(f"{DATA_DIR}/all_playoff_dfs.csv")
    if all_playoff_dfs.empty:
        return pd.DataFrame(), pd.DataFrame()
    print(f"All playoff DataFrame:\n{all_playoff_dfs}")

    # Analyze playoff chances for teams starting 0-1 through 0-7
    playoff_chances = []
    for league_config in leagues:
        league_id = league_config['league_id']
        espn_s2 = league_config['espn_s2']
        swid = league_config['swid']
        league_name = league_config['name']

        for year in years:
            try:
                league = League(league_id=league_id, year=year, espn_s2=espn_s2, swid=swid)
            except Exception as e:
                # print(f"Error initializing league {league_name} for year {year}: {e}")
                continue

            # Filter the playoff DataFrame for the current year and league
            year_playoff_df = all_playoff_dfs[
                (all_playoff_dfs['Year'] == year) & (all_playoff_dfs['League'] == league_name)
            ]

            # Combine Team 1 and Team 2 columns to get all teams that participated in the playoffs
            playoff_teams = pd.concat([year_playoff_df['Team 1'], year_playoff_df['Team 2']])
            playoff_teams = playoff_teams[playoff_teams != 'Bye'].unique()

            print(f"Playoff teams for league: {league_name}, year {year}: {playoff_teams}")

            # Loop through undefeated records from 1-0 to 7-0
            for win_record in range(1, 8):
                # Get teams with the current undefeated record
                undefeated_teams = [
                    team.team_name
                    for team in league.teams
                    if team.outcomes[:win_record].count('W') == win_record
                ]
                print(f"Undefeated teams with record {win_record}-0 for league: {league_name}, year {year}: {undefeated_teams}")

                # Stop searching if no teams match the current undefeated record
                if not undefeated_teams:
                    break

                # Calculate the percentage of undefeated teams that made the playoffs
                made_playoffs = [team for team in undefeated_teams if team in playoff_teams]
                playoff_percentage = len(made_playoffs) / len(undefeated_teams) * 100

                # Append the result to the playoff chances list
                playoff_chances.append({
                    "League": league_name,
                    "Year": year,
                    "Undefeated Record": f"{win_record}-0",
                    "Total Teams": len(undefeated_teams),
                    "Made Playoffs": len(made_playoffs),
                    "Playoff Percentage": playoff_percentage
                })

    # Convert the playoff chances list to a DataFrame
    playoff_chances_df = pd.DataFrame(playoff_chances)

    # Aggregate totals, made playoffs, and playoff percentage by year and league
    aggregated_df = playoff_chances_df.groupby(['Undefeated Record']).agg({
        'Total Teams': 'sum',
        'Made Playoffs': 'sum'
    }).reset_index()

    # Calculate the aggregated playoff percentage
    aggregated_df['Playoff Percentage'] = (aggregated_df['Made Playoffs'] / aggregated_df['Total Teams']) * 100

    # Return the final DataFrames
    return aggregated_df, playoff_chances_df

def all_record_chances(leagues, years):
    def create_playoff_df(leagues, years):
        
        # Initialize an empty list to store all playoff data
        combined_playoff_dfs = []

        # Loop through each league
        for league_config in leagues:
            league_id = league_config['league_id']
            espn_s2 = league_config['espn_s2']
            swid = league_config['swid']
            league_name = league_config['name']

            for year in years:
                print(f"Processing league: {league_name}, year: {year}")

                # Instantiate the league object for the current year
                try:
                    league = League(league_id=league_id, year=year, espn_s2=espn_s2, swid=swid)
                except Exception as e:
                    # print(f"Error initializing league {league_name} for year {year}: {e}")
                    continue

                # Get the league name and construct the file path
                file_path = f"{LEAGUES_DIR}/{league_name} {year}.xlsx"

                # Check if the file exists
                if not os.path.exists(file_path):
                    # print(f"File not found for year {year}: {file_path}. Skipping this year.")
                    continue

                try:
                    # Load the workbook
                    workbook = load_workbook(file_path, read_only=True)

                    # Check if "Playoff Results" sheet exists
                    if "Playoff Results" not in workbook.sheetnames:
                        # print(f"Skipping {file_path}: 'Playoff Results' sheet not found.")
                        continue

                    # Read the Playoff Results sheet into a DataFrame
                    playoff_df = pd.read_excel(file_path, sheet_name="Playoff Results")
                    playoff_df['Year'] = year  # Add year for identification
                    playoff_df['League'] = league_name  # Add league name for identification
                    # basename only - an absolute path here would be machine-specific
                    # and would break consumers that match on this column
                    playoff_df['File Name'] = os.path.basename(file_path)
                    combined_playoff_dfs.append(playoff_df)

                    print(f"Processed {file_path}: 'Playoff Results' sheet loaded.")

                except Exception as e:
                    print(f"Error processing {file_path}: {e}")

        # Combine all playoff DataFrames into one
        if not combined_playoff_dfs:
            # print("No playoff data found.")
            return pd.DataFrame(), pd.DataFrame()  # Return empty DataFrames if no data is found

        all_playoff_dfs = pd.concat(combined_playoff_dfs, ignore_index=True)
        all_playoff_dfs.to_csv(f"{DATA_DIR}/all_playoff_dfs.csv", index=False)
        print(f"Combined playoff DataFrame:\n{all_playoff_dfs}")
        return all_playoff_dfs
    # all_playoff_dfs = create_playoff_df(leagues, years)
    all_playoff_dfs = pd.read_csv(f"{DATA_DIR}/all_playoff_dfs.csv")
    if all_playoff_dfs.empty:
        return pd.DataFrame(), pd.DataFrame()
    print(f"All playoff DataFrame:\n{all_playoff_dfs}")

    
    return combined_aggregated_df, combined_playoff_df
# Example usage
league_id = 1118513122
espn_s2 = CRED["louie_s2_pages"]
espn_s2 = CRED["louie_s2"]
swid = CRED["louie_swid"]


louie_s2 = CRED["louie_s2"]
prahlad_s2 = CRED["prahlad_s2"]
la_s2 = CRED["la_s2"]
hannah_s2 = CRED["hannah_s2"]
ava_s2 = CRED["ava_s2"]
matt_s2 = CRED["matt_s2"]
elle_s2 = CRED["elle_s2"]
# List of league configurations
year = 2025
leagues = [
    # Pennoni Younglings
    {"league_id": 310334683, "year": year, "espn_s2": louie_s2, "swid": CRED["louie_swid"], "name": "Pennoni Younglings"},
    # Family League
    {"league_id": 996930954, "year": year, "espn_s2": louie_s2, "swid": CRED["louie_swid"], "name": "Family League"},
    # EBC League
    {"league_id": 1118513122, "year": year, "espn_s2": louie_s2, "swid": CRED["louie_swid"], "name": "EBC League"},
    # Pennoni Transportation
    {"league_id": 1339704102, "year": year, "espn_s2": prahlad_s2, "swid": CRED["prahlad_swid"], "name": "0755 Fantasy Football"},
    # Game of Yards
    {"league_id": 1781851, "year": year, "espn_s2": prahlad_s2, "swid": CRED["prahlad_swid"], "name": "Game of Yards!"},
    # Brown Munde
    {"league_id": 367134149, "year": year, "espn_s2": prahlad_s2, "swid": CRED["prahlad_swid"], "name": "Brown Munde"},
    # Other Prahlad League
    {"league_id":1242265374, "year":year, "espn_s2":CRED["turf_s2"], "swid":CRED["prahlad_swid"], "name": "Brown Munde"},
    # Las League
    {"league_id": 1049459, "year": year, "espn_s2": la_s2, "swid": CRED["la_swid"], "name": "THE BEST OF THE BEST"},
    # Hannahs League
    {"league_id": 1399036372, "year": year, "espn_s2": hannah_s2, "swid": CRED["hannah_swid"], "name": "Hannahs League"},
    # Avas League
    {"league_id": 417131856, "year": year, "espn_s2": ava_s2, "swid": CRED["ava_swid"], "name": "Avas League"},
    # Matts League
    {"league_id": 261375772, "year": year, "espn_s2": matt_s2, "swid": CRED["matt_swid"], "name": "Matts League"},
    # Elles League
    {"league_id": 1259693145, "year": year, "espn_s2": elle_s2, "swid": CRED["elle_swid"], "name": "Matts League"},
]

years = [2018, 2019, 2020, 2021, 2022, 2023, 2024]
league = League(league_id=1118513122, year=year, espn_s2=espn_s2, swid=CRED["louie_swid"])

aggregated_df, playoff_chances_df = winless_record_chances(leagues, years)
print(playoff_chances_df)
print()
print(aggregated_df)

aggregated_df.to_csv(f"{DATA_DIR}/playoff_chances_winless.csv", index=False)
aggregated_df, playoff_chances_df = undefeated_record_chances(leagues, years)
print(playoff_chances_df)
print()
print(aggregated_df)
aggregated_df.to_csv(f"{DATA_DIR}/playoff_chances_undefeated.csv", index=False)
# playoff_chances_df.to_csv(f"{DATA_DIR}/all_undefeated_playoffs.csv", index=False)
