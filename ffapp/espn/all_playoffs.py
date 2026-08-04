import os as _os, sys as _sys
_d = _os.path.dirname(_os.path.abspath(__file__))
while _d != _os.path.dirname(_d) and not _os.path.exists(_os.path.join(_d, 'paths.py')):
    _d = _os.path.dirname(_d)
_sys.path.insert(0, _d)
from credentials import CRED
import pandas as pd
import os
from openpyxl import load_workbook
from paths import DATA_DIR, LEAGUES_DIR

def create_playoff_df(leagues, years, csv_path=f"{DATA_DIR}/all_playoff_dfs.csv"):
    # imported here, not at module scope, so rebuild_from_workbooks() below
    # stays usable without espn_api or credentials
    from espn_api.football import League
    # Initialize an empty list to store all playoff data
    combined_playoff_dfs = []

    # Loop through each league
    for league_config in leagues:
        league_id = league_config['league_id']
        espn_s2 = league_config['espn_s2']
        swid = league_config['swid']
        league_name = league_config['name']
        if league_name == "Family League":
            league_name = "Family Fantasy"

        for year in years:
            print(f"Processing league: {league_name}, year: {year}")

            # Instantiate the league object for the current year
            try:
                league = League(league_id=league_id, year=year, espn_s2=espn_s2, swid=swid)
                print(league.settings)
                print(league.current_week)
                print(league_name)
            except Exception:
                print(f"Error initializing league {league_name} for year {year}")
                continue

            file_path = f"{LEAGUES_DIR}/{league_name} {year}.xlsx"

            # Check if the file exists
            if not os.path.exists(file_path):
                continue

            try:
                workbook = load_workbook(file_path, read_only=True)

                if "Playoff Results" not in workbook.sheetnames:
                    continue

                # Read the Playoff Results sheet into a DataFrame
                playoff_df = pd.read_excel(file_path, sheet_name="Playoff Results")
                playoff_df['Year'] = year
                playoff_df['League'] = league_name
                playoff_df['File Name'] = file_path
                combined_playoff_dfs.append(playoff_df)

                print(f"Processed {file_path}: 'Playoff Results' sheet loaded.")

            except Exception as e:
                print(f"Error processing {file_path}: {e}")
            print()

    if not combined_playoff_dfs:
        return pd.DataFrame()

    # Combine all newly gathered playoff data
    new_playoff_dfs = pd.concat(combined_playoff_dfs, ignore_index=True)

    # If the CSV already exists, load it and append
    if os.path.exists(csv_path):
        existing_df = pd.read_csv(csv_path)
        # Avoid duplicating if same league+year already exists
        combined_df = pd.concat([existing_df, new_playoff_dfs], ignore_index=True).drop_duplicates()
        # Drop duplicates based on key columns
        combined_df = combined_df.drop_duplicates(
            subset=["Round", "Winner", "Year", "League", "File Name"]
        )

        # Save back to CSV
        combined_df.to_csv(csv_path, index=False)
    else:
        combined_df = new_playoff_dfs

    # Save back to CSV
    combined_df.to_csv(csv_path, index=False)

    print(f"Updated playoff data saved to {csv_path}")
    return combined_df

# louie_s2 = CRED["louie_s2"]
# prahlad_s2 = CRED["prahlad_s2"]
# la_s2 = CRED["la_s2"]
# hannah_s2 = CRED["hannah_s2"]
# ava_s2 = CRED["ava_s2"]
# matt_s2 = CRED["matt_s2"]
# elle_s2 = CRED["elle_s2"]
# # List of league configurations
# year = 2023
# leagues = [
#     # Pennoni Younglings
#     {"league_id": 310334683, "year": year, "espn_s2": louie_s2, "swid": CRED["louie_swid"], "name": "Pennoni Younglings"},
#     # Family League
#     {"league_id": 996930954, "year": year, "espn_s2": louie_s2, "swid": CRED["louie_swid"], "name": "Family League"},
#     # EBC League
#     {"league_id": 1118513122, "year": year, "espn_s2": louie_s2, "swid": CRED["louie_swid"], "name": "EBC League"},
#     # Pennoni Transportation
#     {"league_id": 1339704102, "year": year, "espn_s2": prahlad_s2, "swid": CRED["prahlad_swid"], "name": "0755 Fantasy Football"},
#     # Game of Yards
#     {"league_id": 1781851, "year": year, "espn_s2": prahlad_s2, "swid": CRED["prahlad_swid"], "name": "Game of Yards!"},
#     # Brown Munde
#     {"league_id": 367134149, "year": year, "espn_s2": prahlad_s2, "swid": CRED["prahlad_swid"], "name": "Brown Munde"},
#     # Turf on Grade League
#     {"league_id":1242265374, "year":year, "espn_s2":CRED["turf_s2"], "swid":CRED["prahlad_swid"], "name": "Turf On Grade 2.0"},
#     # Las League
#     {"league_id": 1049459, "year": year, "espn_s2": la_s2, "swid": CRED["la_swid"], "name": "THE BEST OF THE BEST"},
#     # Hannahs League
#     {"league_id": 1399036372, "year": year, "espn_s2": hannah_s2, "swid": CRED["hannah_swid"], "name": "Hannahs League"},
#     # Avas League
#     {"league_id": 417131856, "year": year, "espn_s2": ava_s2, "swid": CRED["ava_swid"], "name": "Avas League"},
#     # Matts League
#     {"league_id": 261375772, "year": year, "espn_s2": matt_s2, "swid": CRED["matt_swid"], "name": "Matts League"},
#     # Elles League
#     {"league_id": 1259693145, "year": year, "espn_s2": elle_s2, "swid": CRED["elle_swid"], "name": "Matts League"},
# ]

# years = [2019, 2020, 2021, 2022, 2023, 2024]
# # years = [2023]
# create_playoff_df(leagues, years)

def rebuild_from_workbooks(csv_path=None):
    """Rebuild the playoff aggregate purely from the league workbooks.

    `create_playoff_df` above instantiates an ESPN League for every league-year
    before reading the workbook - and then never uses it - so refreshing this
    aggregate needed live credentials it does not actually depend on. The
    workbooks are written by the weekly update and are the real source of truth,
    so this reads them directly and can run any time, offline.

    Rebuilds rather than appends: the old append-and-dedupe path left the file
    permanently stale for any season whose workbook changed after its first pull
    (which is why 2025 playoff results were missing everywhere while sitting
    right there in the workbooks).
    """
    import glob
    import re
    csv_path = csv_path or f"{DATA_DIR}/all_playoff_dfs.csv"
    frames = []
    for path in sorted(glob.glob(f"{LEAGUES_DIR}/*.xlsx")):
        base = os.path.splitext(os.path.basename(path))[0]
        m = re.match(r'^(.+) (\d{4})$', base)
        if not m:
            continue
        league_name, year = m.group(1), int(m.group(2))
        try:
            workbook = load_workbook(path, read_only=True)
            if "Playoff Results" not in workbook.sheetnames:
                workbook.close()
                continue
            workbook.close()
            df = pd.read_excel(path, sheet_name="Playoff Results")
        except Exception as e:
            print(f"  skipped {base}: {e}")
            continue
        if df.empty:
            continue
        df = df.loc[:, ~df.columns.astype(str).str.startswith('Unnamed')]
        df['Year'] = year
        df['League'] = league_name
        df['File Name'] = path
        frames.append(df)

    if not frames:
        print("no playoff sheets found")
        return pd.DataFrame()
    out = pd.concat(frames, ignore_index=True)
    out.to_csv(csv_path, index=False)
    print(f"rebuilt {csv_path}: {len(out)} rows, "
          f"{out['League'].nunique()} leagues, years {sorted(out['Year'].unique())}")
    return out
