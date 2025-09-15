from espn_api.football import League
import pandas as pd
import os
from openpyxl import load_workbook

def analyze_playoff_chances(league_id, espn_s2, swid, years):
    """
    Analyzes the percentage chance of making the playoffs for teams starting winless at 0-1 through 0-7.

    Parameters:
    - league_id (int): The league ID.
    - espn_s2 (str): ESPN S2 authentication token.
    - swid (str): SWID authentication token.
    - years (list): List of years to analyze.

    Returns:
    - pd.DataFrame: A DataFrame summarizing the playoff chances for each winless record.
    """
    # Initialize an empty list to store all playoff data
    combined_playoff_dfs = []

    # Loop through each year
    for year in years:
        print(f"Processing year: {year}")

        # Instantiate the league object for the current year
        league = League(league_id=league_id, year=year, espn_s2=espn_s2, swid=swid)

        # Get the league name and construct the file path
        league_name = league.settings.name.replace(" 22/23", "")
        file_path = f"leagues/{league_name} {year}.xlsx"

        # Check if the file exists
        if not os.path.exists(file_path):
            print(f"File not found for year {year}: {file_path}. Skipping this year.")
            continue

        try:
            # Load the workbook
            workbook = load_workbook(file_path, read_only=True)

            # Check if "Playoff Results" sheet exists
            if "Playoff Results" not in workbook.sheetnames:
                print(f"Skipping {file_path}: 'Playoff Results' sheet not found.")
                continue

            # Read the Playoff Results sheet into a DataFrame
            playoff_df = pd.read_excel(file_path, sheet_name="Playoff Results")
            playoff_df['Year'] = year  # Add year for identification
            playoff_df['File Name'] = file_path  # Add file name for identification
            combined_playoff_dfs.append(playoff_df)

            print(f"Processed {file_path}: 'Playoff Results' sheet loaded.")

        except Exception as e:
            print(f"Error processing {file_path}: {e}")

    # Combine all playoff DataFrames into one
    if not combined_playoff_dfs:
        print("No playoff data found.")
        return pd.DataFrame()  # Return an empty DataFrame if no data is found

    all_playoff_dfs = pd.concat(combined_playoff_dfs, ignore_index=True)
    print(f"Combined playoff DataFrame:\n{all_playoff_dfs}")

    # Analyze playoff chances for teams starting 0-1 through 0-7
    playoff_chances = []
    for year in years:
        # Filter the playoff DataFrame for the current year
        year_playoff_df = all_playoff_dfs[all_playoff_dfs['Year'] == year]

        # Get the list of teams that made the playoffs
        playoff_teams = year_playoff_df['Team'].unique()

        # Loop through winless records from 0-1 to 0-7
        for winless_record in range(1, 8):
            # Get teams with the current winless record
            winless_teams = league.teams
            winless_teams = [
                team.team_name
                for team in winless_teams
                if team.scores[:winless_record].count(0) == winless_record
            ]

            # Calculate the percentage of winless teams that made the playoffs
            if winless_teams:
                made_playoffs = [team for team in winless_teams if team in playoff_teams]
                playoff_percentage = len(made_playoffs) / len(winless_teams) * 100
            else:
                playoff_percentage = 0.0

            # Append the result to the playoff chances list
            playoff_chances.append({
                "Year": year,
                "Winless Record": f"0-{winless_record}",
                "Total Teams": len(winless_teams),
                "Made Playoffs": len(made_playoffs),
                "Playoff Percentage": playoff_percentage
            })

    # Convert the playoff chances list to a DataFrame
    playoff_chances_df = pd.DataFrame(playoff_chances)

    # Return the final DataFrame
    return playoff_chances_df


# Example usage
league_id = 1118513122
espn_s2 = 'AEB%2Bzu7FGxYPXt8rgNkQWTV8c4yxT2T3KNZZVkZUVKh9TOdH7iUalV08hSloqYJ5dDtxZVK6d4WC503CH3mH0UkNCPOgbTXYz44W3IJtXsplT%2BLoqNYCU8T7W1HU%2Fgh4PnasvHIkDZgTZFWkUFhcLA0eLkwH8AvYe2%2FCIlhdk7%2FdMeiM0ijsS8vhSYYB8LUhSrB0kuTXE2v85gSIrJQSbs3mPvP5p6pFr3w2OxWicVi9pe8p3eVDhSOLiPMYrPgpuL%2FLBZIGHxhKz5lzGRSL2uTA'
espn_s2 = "AECL47AORj8oAbgOmiQidZQsoAJ6I8ziOrC8Jw0W2M0QwSjYsyUkzobZA0CZfGBYrKf0a%2B%2B3%2Fflv6rFCZvb3%2FWo%2FfKVU4JXm9UyLsY9uIRAF4o9TuISaQjoc13SbsqMiLyaf5kR4ZwDcNr8uUxDwamEyuec5yqs07zsvy0VrOQo6NTxylWXkwABFfNVAdyqDI%2BQoQtoetdSah0eYfMdmSIBkGnxN0R0z5080zBAuY9yCm%2Fav49lUfGA7cqGyWoIky8pE3vB%2Fng%2F49JvTerFjJfzC"
swid = '{4656A2AD-A939-460B-96A2-ADA939760B8B}'
years = [2021, 2022, 2023, 2024]
# league = League(league_id=1118513122, year=year, espn_s2=espn_s2, swid='{4656A2AD-A939-460B-96A2-ADA939760B8B}')

playoff_chances_df = analyze_playoff_chances(league_id, espn_s2, swid, years)
print(playoff_chances_df)