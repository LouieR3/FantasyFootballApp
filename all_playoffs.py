from credentials import CRED
from espn_api.football import League
import pandas as pd
import os
from openpyxl import load_workbook

def create_playoff_df(leagues, years, csv_path="all_playoff_dfs.csv"):
    # Initialize an empty list to store all playoff data
    combined_playoff_dfs = []

    # Loop through each league
    for league_config in leagues:
        league_id = league_config['league_id']
        espn_s2 = league_config['espn_s2']
        swid = league_config['swid']
        league_name = league_config['name']
        if league_name == "Family League":
            league_name = "Family Fantasy"

        for year in years:
            print(f"Processing league: {league_name}, year: {year}")

            # Instantiate the league object for the current year
            try:
                league = League(league_id=league_id, year=year, espn_s2=espn_s2, swid=swid)
                print(league.settings)
                print(league.current_week)
                print(league_name)
            except Exception:
                print(f"Error initializing league {league_name} for year {year}")
                continue

            file_path = f"leagues/{league_name} {year}.xlsx"

            # Check if the file exists
            if not os.path.exists(file_path):
                continue

            try:
                workbook = load_workbook(file_path, read_only=True)

                if "Playoff Results" not in workbook.sheetnames:
                    continue

                # Read the Playoff Results sheet into a DataFrame
                playoff_df = pd.read_excel(file_path, sheet_name="Playoff Results")
                playoff_df['Year'] = year
                playoff_df['League'] = league_name
                playoff_df['File Name'] = file_path
                combined_playoff_dfs.append(playoff_df)

                print(f"Processed {file_path}: 'Playoff Results' sheet loaded.")

            except Exception as e:
                print(f"Error processing {file_path}: {e}")
            print()

    if not combined_playoff_dfs:
        return pd.DataFrame()

    # Combine all newly gathered playoff data
    new_playoff_dfs = pd.concat(combined_playoff_dfs, ignore_index=True)

    # If the CSV already exists, load it and append
    if os.path.exists(csv_path):
        existing_df = pd.read_csv(csv_path)
        # Avoid duplicating if same league+year already exists
        combined_df = pd.concat([existing_df, new_playoff_dfs], ignore_index=True).drop_duplicates()
        # Drop duplicates based on key columns
        combined_df = combined_df.drop_duplicates(
            subset=["Round", "Winner", "Year", "League", "File Name"]
        )

        # Save back to CSV
        combined_df.to_csv(csv_path, index=False)
    else:
        combined_df = new_playoff_dfs

    # Save back to CSV
    combined_df.to_csv(csv_path, index=False)

    print(f"Updated playoff data saved to {csv_path}")
    return combined_df

# louie_s2 = CRED["louie_s2"]
# prahlad_s2 = CRED["prahlad_s2"]
# la_s2 = CRED["la_s2"]
# hannah_s2 = CRED["hannah_s2"]
# ava_s2 = CRED["ava_s2"]
# matt_s2 = CRED["matt_s2"]
# elle_s2 = CRED["elle_s2"]
# # List of league configurations
# year = 2023
# leagues = [
#     # Pennoni Younglings
#     {"league_id": 310334683, "year": year, "espn_s2": louie_s2, "swid": CRED["louie_swid"], "name": "Pennoni Younglings"},
#     # Family League
#     {"league_id": 996930954, "year": year, "espn_s2": louie_s2, "swid": CRED["louie_swid"], "name": "Family League"},
#     # EBC League
#     {"league_id": 1118513122, "year": year, "espn_s2": louie_s2, "swid": CRED["louie_swid"], "name": "EBC League"},
#     # Pennoni Transportation
#     {"league_id": 1339704102, "year": year, "espn_s2": prahlad_s2, "swid": CRED["prahlad_swid"], "name": "0755 Fantasy Football"},
#     # Game of Yards
#     {"league_id": 1781851, "year": year, "espn_s2": prahlad_s2, "swid": CRED["prahlad_swid"], "name": "Game of Yards!"},
#     # Brown Munde
#     {"league_id": 367134149, "year": year, "espn_s2": prahlad_s2, "swid": CRED["prahlad_swid"], "name": "Brown Munde"},
#     # Turf on Grade League
#     {"league_id":1242265374, "year":year, "espn_s2":CRED["turf_s2"], "swid":CRED["prahlad_swid"], "name": "Turf On Grade 2.0"},
#     # Las League
#     {"league_id": 1049459, "year": year, "espn_s2": la_s2, "swid": CRED["la_swid"], "name": "THE BEST OF THE BEST"},
#     # Hannahs League
#     {"league_id": 1399036372, "year": year, "espn_s2": hannah_s2, "swid": CRED["hannah_swid"], "name": "Hannahs League"},
#     # Avas League
#     {"league_id": 417131856, "year": year, "espn_s2": ava_s2, "swid": CRED["ava_swid"], "name": "Avas League"},
#     # Matts League
#     {"league_id": 261375772, "year": year, "espn_s2": matt_s2, "swid": CRED["matt_swid"], "name": "Matts League"},
#     # Elles League
#     {"league_id": 1259693145, "year": year, "espn_s2": elle_s2, "swid": CRED["elle_swid"], "name": "Matts League"},
# ]

# years = [2019, 2020, 2021, 2022, 2023, 2024]
# # years = [2023]
# create_playoff_df(leagues, years)