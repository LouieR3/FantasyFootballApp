def app():
    import pandas as pd
    from operator import itemgetter
    import glob
    import streamlit as st
    import pandas as pd
    from espn_api.football import League
    import pandas as pd
    from openpyxl import load_workbook
    import time

    pd.options.mode.chained_assignment = None
    st.header('League Analysis Across All Leagues')
    st.subheader('In looking at regular season and playoff data across all leagues, how well does LPI predict winners? How about seeds, records, or total points? How often do top seeds win the championship? How often do lower seeds win the championship?')
    st.divider()
    # Initialize an empty list to store all playoff_dfs
    combined_playoff_dfs = []

    # Get all Excel files in the current directory
    xlsx_files = glob.glob("*.xlsx")

    # Loop through each file
    for file in xlsx_files:
        try:
            # Load the workbook
            workbook = load_workbook(file, read_only=True)

            # Check if "Playoff Results" sheet exists
            if "Playoff Results" not in workbook.sheetnames:
                # print(f"Skipping {file}: 'Playoff Results' sheet not found.")
                continue

            # Read the Playoff Results sheet into a DataFrame
            playoff_df = pd.read_excel(file, sheet_name="Playoff Results")
            playoff_df['File Name'] = file  # Add file name for identification
            combined_playoff_dfs.append(playoff_df)

            # print(f"Processed {file}: 'Playoff Results' sheet loaded.")

        except Exception as e:
            print(f"Error processing {file}: {e}")

    # Combine all DataFrames into one
    if combined_playoff_dfs:
        all_playoff_dfs = pd.concat(combined_playoff_dfs, ignore_index=True)
        # print("Combined all playoff data successfully.")
    else:
        print("No valid playoff data found.")


    # Combine all DataFrames into one
    all_playoff_dfs = pd.concat(combined_playoff_dfs, ignore_index=True)
    # print(all_playoff_dfs)

    # Ensure LPI columns are numeric
    all_playoff_dfs['LPI 1'] = pd.to_numeric(all_playoff_dfs['LPI 1'], errors='coerce')
    all_playoff_dfs['LPI 2'] = pd.to_numeric(all_playoff_dfs['LPI 2'], errors='coerce')

    # Ensure Seed columns are numeric (integers)
    all_playoff_dfs['Seed 1'] = pd.to_numeric(all_playoff_dfs['Seed 1'], errors='coerce', downcast='integer')
    all_playoff_dfs['Seed 2'] = pd.to_numeric(all_playoff_dfs['Seed 2'], errors='coerce', downcast='integer')

    # Ensure Total Points columns are numeric (floats)
    all_playoff_dfs['Total Points 1'] = pd.to_numeric(all_playoff_dfs['Total Points 1'], errors='coerce')
    all_playoff_dfs['Total Points 2'] = pd.to_numeric(all_playoff_dfs['Total Points 2'], errors='coerce')

    # Check for NaN values and fill with default values
    if all_playoff_dfs[['LPI 1', 'LPI 2', 'Seed 1', 'Seed 2', 
                        'Total Points 1', 'Total Points 2']].isna().any().any():
        # print("Warning: Some values are missing or invalid.")
        all_playoff_dfs['LPI 1'] = all_playoff_dfs['LPI 1'].fillna(0)
        all_playoff_dfs['LPI 2'] = all_playoff_dfs['LPI 2'].fillna(0)
        all_playoff_dfs['Seed 1'] = all_playoff_dfs['Seed 1'].fillna(0).astype(int)  # Default seed to 0 and ensure integer
        all_playoff_dfs['Seed 2'] = all_playoff_dfs['Seed 2'].fillna(0).astype(int)
        all_playoff_dfs['Total Points 1'] = all_playoff_dfs['Total Points 1'].fillna(0.0)
        all_playoff_dfs['Total Points 2'] = all_playoff_dfs['Total Points 2'].fillna(0.0)


    # Identify and skip bye games
    all_playoff_dfs['Is Bye'] = all_playoff_dfs['Team 2'] == 'Bye'
    # print(all_playoff_dfs)
    # print(all_playoff_dfs[all_playoff_dfs['Is Bye'] == False])

    def correct_percentages(all_playoff_dfs):
        no_byes = all_playoff_dfs[all_playoff_dfs['Is Bye'] == False]
        # --------------------------------------------------------------------------------------------
        # Compare LPI
        no_byes['LPI Winner'] = (
            ((no_byes['LPI 1'] > no_byes['LPI 2']) & 
            (no_byes['Winner'] == no_byes['Team 1'])) |
            ((no_byes['LPI 2'] > no_byes['LPI 1']) & 
            (no_byes['Winner'] == no_byes['Team 2']))
        )

        # Count LPI correct predictions
        lpi_correct = no_byes['LPI Winner'].sum()
        total_games = len(no_byes)
        lpi_accuracy = lpi_correct / total_games * 100
        print(f"LPI was correct in predicting the winner in {lpi_correct}/{total_games} games ({lpi_accuracy:.2f}%)")
        st.write(f"LPI was correct in predicting the winner in {lpi_correct}/{total_games} games ({lpi_accuracy:.2f}%)")
        # --------------------------------------------------------------------------------------------

        # --------------------------------------------------------------------------------------------
        # Compare Total Points
        no_byes['Total Points Winner'] = (
            ((no_byes['Total Points 1'] > no_byes['Total Points 2']) & 
            (no_byes['Winner'] == no_byes['Team 1'])) |
            ((no_byes['Total Points 2'] > no_byes['Total Points 1']) & 
            (no_byes['Winner'] == no_byes['Team 2']))
        )

        # Count Total Points correct predictions
        total_points_correct = no_byes['Total Points Winner'].sum()
        total_games = len(no_byes)
        total_points_accuracy = total_points_correct / total_games * 100

        print(f"Total Points was correct in predicting the winner in {total_points_correct}/{total_games} games ({total_points_accuracy:.2f}%)")
        st.write(f"Total Points was correct in predicting the winner in {total_points_correct}/{total_games} games ({total_points_accuracy:.2f}%)")
        # --------------------------------------------------------------------------------------------

        # --------------------------------------------------------------------------------------------
        # Compare Seed
        no_byes['Seed Winner'] = (
            ((no_byes['Seed 1'] > no_byes['Seed 2']) & 
                (no_byes['Winner'] == no_byes['Team 1'])) |
            ((no_byes['Seed 2'] > no_byes['Seed 1']) & 
                (no_byes['Winner'] == no_byes['Team 2']))
        )

        # Count Total Points correct predictions
        total_points_correct = no_byes['Total Points Winner'].sum()
        total_games = len(no_byes)
        total_points_accuracy = total_points_correct / total_games * 100

        print(f"Seed was correct in predicting the winner in {total_points_correct}/{total_games} games ({total_points_accuracy:.2f}%)")
        st.write(f"Seed was correct in predicting the winner in {total_points_correct}/{total_games} games ({total_points_accuracy:.2f}%)")
        # --------------------------------------------------------------------------------------------

        # --------------------------------------------------------------------------------------------
        # Compare Record
        no_byes['Wins 1'] = no_byes['Record 1'].str.split('-').str[0].astype(int)
        no_byes['Wins 2'] = no_byes['Record 2'].str.split('-').str[0].astype(int)

        # Create a sub-DataFrame excluding cases where the wins are the same
        no_byes_different_wins = no_byes[no_byes['Wins 1'] != no_byes['Wins 2']]

        # Add a column to determine if the team with more wins won
        no_byes_different_wins['More Wins Winner'] = (
            ((no_byes_different_wins['Wins 1'] > no_byes_different_wins['Wins 2']) & 
            (no_byes_different_wins['Winner'] == no_byes_different_wins['Team 1'])) |
            ((no_byes_different_wins['Wins 2'] > no_byes_different_wins['Wins 1']) & 
            (no_byes_different_wins['Winner'] == no_byes_different_wins['Team 2']))
        )

        # Count Total Points correct predictions
        total_points_correct = no_byes_different_wins['More Wins Winner'].sum()
        total_games = len(no_byes_different_wins)
        total_points_accuracy = total_points_correct / total_games * 100

        print(f"Record was correct in predicting the winner in {total_points_correct}/{total_games} games ({total_points_accuracy:.2f}%)")
        st.write(f"Record was correct in predicting the winner in {total_points_correct}/{total_games} games ({total_points_accuracy:.2f}%)")
        print()
        st.divider()
        # print(no_byes_different_wins[["Team 1", "Seed 1", "Score 1", "LPI 1", "Record 1", "Team 2", "Seed 2", "Score 2", "LPI 2", "Record 2", "More Wins Winner"]])
        # print()
        # --------------------------------------------------------------------------------------------

        # Calculate win rate for teams with negative LPI
        negative_lpi_games = no_byes[(no_byes['LPI 1'] < 0) | (no_byes['LPI 2'] < 0)]
        negative_lpi_wins = (
            ((negative_lpi_games['LPI 1'] < 0) & (negative_lpi_games['Winner'] == negative_lpi_games['Team 1'])) |
            ((negative_lpi_games['LPI 2'] < 0) & (negative_lpi_games['Winner'] == negative_lpi_games['Team 2']))
        ).sum()
        negative_lpi_total = len(negative_lpi_games)
        negative_lpi_accuracy = negative_lpi_wins / negative_lpi_total * 100
        print(f"Teams with negative LPI won {negative_lpi_wins}/{negative_lpi_total} games ({negative_lpi_accuracy:.2f}%)")
        st.write(f"Teams with negative LPI won {negative_lpi_wins}/{negative_lpi_total} games ({negative_lpi_accuracy:.2f}%)")

        # Calculate win rate for teams with LPI >= 20
        lpi_20_games = no_byes[(no_byes['LPI 1'] >= 20) | (no_byes['LPI 2'] >= 20)]
        lpi_20_wins = (
            ((lpi_20_games['LPI 1'] >= 20) & (lpi_20_games['Winner'] == lpi_20_games['Team 1'])) |
            ((lpi_20_games['LPI 2'] >= 20) & (lpi_20_games['Winner'] == lpi_20_games['Team 2']))
        ).sum()
        lpi_20_total = len(lpi_20_games)
        lpi_20_accuracy = lpi_20_wins / lpi_20_total * 100
        print(f"Teams with LPI >= 20 won {lpi_20_wins}/{lpi_20_total} games ({lpi_20_accuracy:.2f}%)")
        st.write(f"Teams with LPI >= 20 won {lpi_20_wins}/{lpi_20_total} games ({lpi_20_accuracy:.2f}%)")

        # Calculate win rate for teams with LPI >= 40
        lpi_40_games = no_byes[(no_byes['LPI 1'] >= 40) | (no_byes['LPI 2'] >= 40)]
        lpi_40_wins = (
            ((lpi_40_games['LPI 1'] >= 40) & (lpi_40_games['Winner'] == lpi_40_games['Team 1'])) |
            ((lpi_40_games['LPI 2'] >= 40) & (lpi_40_games['Winner'] == lpi_40_games['Team 2']))
        ).sum()
        lpi_40_total = len(lpi_40_games)
        lpi_40_accuracy = lpi_40_wins / lpi_40_total * 100
        print(f"Teams with LPI >= 40 won {lpi_40_wins}/{lpi_40_total} games ({lpi_40_accuracy:.2f}%)")
        st.write(f"Teams with LPI >= 40 won {lpi_40_wins}/{lpi_40_total} games ({lpi_40_accuracy:.2f}%)")

        # Select relevant columns from no_byes and no_byes_different_wins for merging
        # no_byes_merge = no_byes[['Team 1', 'Team 2', 'File Name', 'LPI Winner', 'Total Points Winner', 'Seed Winner']]
        # no_byes_different_wins_merge = no_byes_different_wins[['Team 1', 'Team 2', 'File Name', 'More Wins Winner']]

        # # Merge no_byes data into all_playoff_dfs
        # all_playoff_dfs = all_playoff_dfs.merge(no_byes_merge, on=['Team 1', 'Team 2', 'File Name'], how='left')

        # # Merge no_byes_different_wins data into all_playoff_dfs
        # all_playoff_dfs = all_playoff_dfs.merge(no_byes_different_wins_merge, on=['Team 1', 'Team 2', 'File Name'], how='left')

        # Display the updated DataFrame
        # print(all_playoff_dfs)
        return all_playoff_dfs

    all_playoff_dfs = correct_percentages(all_playoff_dfs)
    # all_playoff_dfs.to_csv("all_playoffs.csv", index=False)
    st.divider()
    def wins_by_seed():
        # --------------------------------------------------------------------------------------------
        # Calculate win counts for each seed
        seed_win_counts_1 = all_playoff_dfs.groupby('Seed 1').apply(
            lambda group: (group['Winner'] == group['Team 1']).sum()
        )
        seed_win_counts_2 = all_playoff_dfs.groupby('Seed 2').apply(
            lambda group: (group['Winner'] == group['Team 2']).sum()
        )

        # Sum the win counts across both seeds
        seed_win_counts = seed_win_counts_1.add(seed_win_counts_2, fill_value=0)

        # Calculate game counts for each seed
        seed_game_counts = all_playoff_dfs['Seed 1'].value_counts().add(
            all_playoff_dfs['Seed 2'].value_counts(), fill_value=0
        )

        # Calculate win rates
        seed_win_rates = (seed_win_counts / seed_game_counts * 100).sort_index()


        # Calculate game counts for each record
        seed_game_counts_1 = all_playoff_dfs['Seed 1'].value_counts()
        seed_game_counts_2 = all_playoff_dfs['Seed 2'].value_counts()
        seed_game_counts = seed_game_counts_1.add(seed_game_counts_2, fill_value=0)

        # Calculate loss counts
        seed_loss_counts = seed_game_counts - seed_win_counts
        total_record = seed_win_counts.astype(int).astype(str) + "-" + seed_loss_counts.astype(int).astype(str)

        # Remove seed 0 (if it exists)
        if 0 in seed_win_rates.index:
            seed_win_rates = seed_win_rates.drop(index=0)
            seed_game_counts = seed_game_counts.drop(index=0)
            total_record = total_record.drop(index=0)

        # Create the result DataFrame
        seed_summary_df = pd.DataFrame({
            "Win Rate": seed_win_rates.round(2).map(lambda x: f"{x:.2f}%"),
            "Total Record": total_record,
            "Total Games": seed_game_counts.astype(int)
        })
        seed_summary_df.index = seed_summary_df.index.map(lambda x: f"Seed {x}")
        # Print results
        print("Win rates by seed:")
        st.write("Win rates by seed:")
        print(seed_summary_df)
        st.dataframe(seed_summary_df)
        st.write()
        print()
        # --------------------------------------------------------------------------------------------

        # --------------------------------------------------------------------------------------------
        # Filter first round games
        total_seasons = all_playoff_dfs['File Name'].nunique()
        # print(all_playoff_dfs[all_playoff_dfs['Round'] == 'Championship'])
        # print(total_seasons)
        # sfd
        first_round_games = all_playoff_dfs[all_playoff_dfs['Round'] == 'Quarter Final']

        # Map winner seeds
        first_round_wins_by_seed = first_round_games['Winner'].map(
        lambda x: all_playoff_dfs.loc[
                (all_playoff_dfs['Team 1'] == x), 'Seed 1'
            ].values[0] if x in all_playoff_dfs['Team 1'].values else 
            all_playoff_dfs.loc[
                (all_playoff_dfs['Team 2'] == x), 'Seed 2'
            ].values[0]
        ).value_counts()

        # Count total first round appearances by seed
        first_round_game_counts = first_round_games['Seed 1'].value_counts().add(
            first_round_games['Seed 2'].value_counts(), fill_value=0
        ).astype(int)

        # Ensure both indices align, and fill missing values with 0
        all_seeds = first_round_game_counts.index.union(first_round_wins_by_seed.index)
        first_round_wins_by_seed = first_round_wins_by_seed.reindex(all_seeds, fill_value=0)
        first_round_game_counts = first_round_game_counts.reindex(all_seeds, fill_value=0)

        # Calculate loss counts
        seed_loss_counts = first_round_game_counts - first_round_wins_by_seed

        # Calculate win probabilities
        first_round_probabilities = (
            (first_round_wins_by_seed / first_round_game_counts * 100)
            .fillna(0)
            .round(2)
            .map(lambda x: f"{x:.2f}%")
        )

        # Ensure Total Record values are properly aligned
        total_record = (
            first_round_wins_by_seed.astype(int).astype(str) + "-" +
            seed_loss_counts.astype(int).astype(str)
        )   

        # Remove seed 0 (if it exists)
        if 0 in seed_win_rates.index:
            seed_win_rates = seed_win_rates.drop(index=0)
            seed_game_counts = seed_game_counts.drop(index=0)
            total_record = total_record.drop(index=0)

        # Combine into a DataFrame
        first_round_summary_df = pd.DataFrame({
            "Win Probability": first_round_probabilities,
            "Total Record": total_record,
            "Total Games": first_round_game_counts
        })
        first_round_summary_df.index = first_round_summary_df.index.map(lambda x: f"Seed {int(x)}")

        # Remove "Seed 0" row if it exists
        if "Seed 0" in first_round_summary_df.index:
            first_round_summary_df = first_round_summary_df.drop(index="Seed 0")

        # Print results
        print("First round win probabilities by seed:")
        print(first_round_summary_df)
        print()
        st.write("First round win probabilities by seed:")
        st.dataframe(first_round_summary_df)
        st.write()
        # --------------------------------------------------------------------------------------------

        # --------------------------------------------------------------------------------------------
        # Filter second round games
        second_round_games = all_playoff_dfs[all_playoff_dfs['Round'] == 'Semi Final']

        # Map winner seeds
        second_round_wins_by_seed = second_round_games['Winner'].map(
            lambda x: all_playoff_dfs.loc[
                (all_playoff_dfs['Team 1'] == x), 'Seed 1'
            ].values[0] if x in all_playoff_dfs['Team 1'].values else 
            all_playoff_dfs.loc[
                (all_playoff_dfs['Team 2'] == x), 'Seed 2'
            ].values[0]
        ).value_counts()

        # Count total second round appearances by seed
        second_round_game_counts = second_round_games['Seed 1'].value_counts().add(
            second_round_games['Seed 2'].value_counts(), fill_value=0
        ).astype(int)

        # Ensure both indices align, and fill missing values with 0
        all_seeds = second_round_game_counts.index.union(second_round_wins_by_seed.index)
        second_round_wins_by_seed = second_round_wins_by_seed.reindex(all_seeds, fill_value=0)
        second_round_game_counts = second_round_game_counts.reindex(all_seeds, fill_value=0)
        # Calculate second round probabilities
        second_round_probabilities = (second_round_wins_by_seed / second_round_game_counts * 100).fillna(0).round(2).map(lambda x: f"{x:.2f}%")

        # Combine into a DataFrame
        second_round_summary_df = pd.DataFrame({
            "Win Probability": second_round_probabilities,
            "Total Games": second_round_game_counts
        })
        second_round_summary_df.index = second_round_summary_df.index.map(lambda x: f"Seed {int(x)}")

        # Remove "Seed 0" row if it exists
        if "Seed 0" in second_round_summary_df.index:
            second_round_summary_df = second_round_summary_df.drop(index="Seed 0")

        # Print results
        print("Semi Final win probabilities by seed:")
        print(second_round_summary_df)
        print()
        st.write("Semi Final win probabilities by seed:")
        st.dataframe(second_round_summary_df)
        st.write()
        # --------------------------------------------------------------------------------------------

        # --------------------------------------------------------------------------------------------
        # Identify championship games
        championship_games = all_playoff_dfs[all_playoff_dfs['Round'] == 'Championship']
        # print(championship_games)
        # print(championship_games[["Team 1", "Seed 1", "Score 1", "LPI 1", "Record 1", "Team 2", "Seed 2", "Score 2", "LPI 2", "Record 2", "File Name"]])
        # print(championship_games[["Team 1", "Seed 1", "Total Points 1", "LPI 1", "Record 1", "Team 2", "Seed 2", "Total Points 2", "LPI 2", "Record 2", "File Name"]])

        # Create a Winner DataFrame by selecting respective columns based on the Winner
        winner_df = championship_games.apply(
            lambda row: pd.Series({
                "Team": row["Team 1"] if row["Winner"] == row["Team 1"] else row["Team 2"],
                "Seed": row["Seed 1"] if row["Winner"] == row["Team 1"] else row["Seed 2"],
                "Total Points": row["Total Points 1"] if row["Winner"] == row["Team 1"] else row["Total Points 2"],
                "LPI": row["LPI 1"] if row["Winner"] == row["Team 1"] else row["LPI 2"],
                "Record": row["Record 1"] if row["Winner"] == row["Team 1"] else row["Record 2"],
                "Score": row["Score 1"] if row["Winner"] == row["Team 1"] else row["Score 2"],
                "File Name": row["File Name"],
            }),
            axis=1
        )

        # Reset the index for the new DataFrame
        winner_df.reset_index(drop=True, inplace=True)

        # Print the Winner DataFrame
        winner_df["File Name"] = winner_df["File Name"].str.replace('.xlsx', '', regex=False)
        print(winner_df)
        st.divider()
        st.write("All Champions:")
        st.dataframe(winner_df)
        st.divider()

        # dfsa
        # Map winner seeds and determine wins by seed
        def get_winner_seed(row):
            if row['Score 1'] > row['Score 2']:
                return row['Seed 1']
            elif row['Score 2'] > row['Score 1']:
                return row['Seed 2']
            else:
                return None  # In case of a tie (if applicable)

        # Apply the function to get the winning seed
        championship_games['Winning Seed'] = championship_games.apply(get_winner_seed, axis=1)

        # Count the number of championships won by each seed (new column)
        championship_wins_by_seed = championship_games['Winning Seed'].value_counts()

        # Count total championship appearances by seed (unchanged)
        championship_game_counts = championship_games['Seed 1'].value_counts().add(
            championship_games['Seed 2'].value_counts(), fill_value=0
        ).astype(int)

        # Calculate the total number of seasons
        total_seasons = championship_games['File Name'].nunique()

        # Ensure both indices align, and fill missing values with 0
        all_seeds = championship_game_counts.index.union(championship_wins_by_seed.index)
        championship_wins_by_seed = championship_wins_by_seed.reindex(all_seeds, fill_value=0)
        championship_game_counts = championship_game_counts.reindex(all_seeds, fill_value=0)

        # Calculate championship probabilities as percentage of total seasons
        championship_probabilities = (championship_wins_by_seed / total_seasons * 100).round(2).map(lambda x: f"{x:.2f}%")

        # Combine into a DataFrame
        championship_summary_df = pd.DataFrame({
            "Championship Probability": championship_probabilities,
            "Total Championships": championship_game_counts,
            "Championships Won": championship_wins_by_seed
        })
        championship_summary_df.index = championship_summary_df.index.map(lambda x: f"Seed {int(x)}")

        # Print results
        print("Out of " + str(total_seasons)+ " seasons analyzed")
        print("Championship win probabilities by seed:")
        print(championship_summary_df)
        print()
        st.divider()
        st.write("Out of " + str(total_seasons)+ " seasons analyzed:")
        st.write("Championship win probabilities by seed:")
        st.dataframe(championship_summary_df)
        st.divider()
        # --------------------------------------------------------------------------------------------

    wins_by_seed()
    st.divider()
    def wins_by_record(all_playoff_dfs):
        print()
        print("WIN RATE BY RECORD:")
        all_playoff_dfs = all_playoff_dfs[all_playoff_dfs['Is Bye'] == False]
        # all_playoff_dfs = all_playoff_dfs[(all_playoff_dfs['Record 1'] == "6-8-0") | (all_playoff_dfs['Record 2'] == "6-8-0")]
        # print(all_playoff_dfs[["Team 1", "Seed 1", "Score 1", "LPI 1", "Record 1", "Team 2", "Seed 2", "Score 2", "LPI 2", "Record 2", "Winner"]])
        # afsd
        all_playoff_dfs['Wins 1'] = all_playoff_dfs['Record 1'].str.split('-').str[0].astype(int)
        all_playoff_dfs['Wins 2'] = all_playoff_dfs['Record 2'].str.split('-').str[0].astype(int)
        all_playoff_dfs['Losses 1'] = all_playoff_dfs['Record 1'].str.split('-').str[1].astype(int)
        all_playoff_dfs['Losses 2'] = all_playoff_dfs['Record 2'].str.split('-').str[1].astype(int)
        # print(all_playoff_dfs)
        # --------------------------------------------------------------------------------------------
        # --------------------------------------------------------------------------------------------
        # Calculate win counts for each record
        record_win_counts_1 = all_playoff_dfs.groupby('Record 1').apply(
            lambda group: (group['Winner'] == group['Team 1']).sum()
        )
        record_win_counts_2 = all_playoff_dfs.groupby('Record 2').apply(
            lambda group: (group['Winner'] == group['Team 2']).sum()
        )

        # Sum the win counts across both records
        record_win_counts = record_win_counts_1.add(record_win_counts_2, fill_value=0)

        # Calculate game counts for each record
        record_game_counts_1 = all_playoff_dfs['Record 1'].value_counts()
        record_game_counts_2 = all_playoff_dfs['Record 2'].value_counts()
        record_game_counts = record_game_counts_1.add(record_game_counts_2, fill_value=0)

        # Calculate win rates
        record_win_rates = (record_win_counts / record_game_counts * 100).round(2)
        # Calculate loss counts
        record_loss_counts = record_game_counts - record_win_counts
        total_record = record_win_counts.astype(int).astype(str) + "-" + record_loss_counts.astype(int).astype(str)

        # Create a new DataFrame with the results
        record_summary_df = pd.DataFrame({
            "Win Rate": record_win_rates.map(lambda x: f"{x:.2f}%"),
            "Total Record": total_record,
            "Total Games": record_game_counts.astype(int)
        })

        # Add a new column for win percentage
        record_summary_df["Win Percentage"] = (
            record_summary_df.index.map(lambda r: int(r.split('-')[0]) / (int(r.split('-')[0]) + int(r.split('-')[1])))
        ).round(3)

        # Filter records that total to 14 games
        record_summary_df["Total Games Played"] = record_summary_df.index.map(lambda r: sum(map(int, r.split('-'))))
        records_14_games = record_summary_df[record_summary_df["Total Games Played"] == 14].copy()

        # # Combine records not totaling 14 games into their corresponding win percentage
        # records_other = record_summary_df[record_summary_df["Total Games Played"] != 14]
        # print(records_14_games)
        # for idx, row in records_other.iterrows():
        #     win_percentage = row["Win Percentage"]
        #     match_idx = records_14_games.index[
        #         records_14_games["Win Percentage"] == win_percentage
        #     ].tolist()

        #     # If a match exists, combine the values
        #     if match_idx:
        #         match_idx = match_idx[0]  # Get the first matching index
        #         # Parse and add Total Record (wins and losses)
        #         existing_wins, existing_losses = map(int, records_14_games.loc[match_idx, "Total Record"].split("-"))
        #         new_wins, new_losses = map(int, row["Total Record"].split("-"))
        #         records_14_games.loc[match_idx, "Total Record"] = f"{existing_wins + new_wins}-{existing_losses + new_losses}"
        #         # Add Total Games
        #         records_14_games.loc[match_idx, "Total Games"] += row["Total Games"]

        # Ensure the index is sorted by the number of wins in ascending order
        record_summary_df.index = pd.Categorical(
            record_summary_df.index, 
            categories=sorted(record_summary_df.index, key=lambda r: int(r.split('-')[0])),
            ordered=True
        )
        record_summary_df = record_summary_df.sort_index()
        record_summary_df.drop(columns=["Total Games Played"], inplace=True)

        # Print results
        print("Win rates by record:")
        # print(record_summary_df)
        # print()
        records_14_games.index = pd.Categorical(
            records_14_games.index, 
            categories=sorted(records_14_games.index, key=lambda r: int(r.split('-')[0])),
            ordered=True
        )
        records_14_games = records_14_games.sort_index()
        records_14_games.drop(columns=["Total Games Played"], inplace=True)
        print(records_14_games)
        print()
        # --------------------------------------------------------------------------------------------

        # --------------------------------------------------------------------------------------------
        # Filter first round games
        total_seasons = all_playoff_dfs['File Name'].nunique()
        # print(all_playoff_dfs[all_playoff_dfs['Round'] == 'Championship'])
        # sfd
        first_round_games = all_playoff_dfs[all_playoff_dfs['Round'] == 'Quarter Final']

        # Initialize counters for the first round wins by record
        first_round_wins_by_record = {}

        # Loop through the first round games and count wins for each record
        for _, game in first_round_games.iterrows():
            # Extract winner and their record (either Record 1 or Record 2)
            winner = game['Winner']
            if winner == game['Team 1']:
                winner_record = game['Record 1']
            else:
                winner_record = game['Record 2']
            
            # Count wins for each record
            if winner_record in first_round_wins_by_record:
                first_round_wins_by_record[winner_record] += 1
            else:
                first_round_wins_by_record[winner_record] = 1

        # Count total appearances of each record in the first round (both Record 1 and Record 2)
        first_round_game_counts = first_round_games['Record 1'].value_counts().add(
            first_round_games['Record 2'].value_counts(), fill_value=0
        ).astype(int)

        # Ensure both indices align, and fill missing values with 0
        all_records = first_round_game_counts.index.union(first_round_wins_by_record.keys())
        first_round_wins_by_record = {record: first_round_wins_by_record.get(record, 0) for record in all_records}
        first_round_game_counts = first_round_game_counts.reindex(all_records, fill_value=0)

        # Calculate first round win probabilities
        first_round_probabilities = (
            pd.Series(first_round_wins_by_record) / first_round_game_counts * 100
        ).fillna(0).round(2).map(lambda x: f"{x:.2f}%")

        # Calculate total record (wins-losses) for each record
        first_round_losses_by_record = first_round_game_counts - pd.Series(first_round_wins_by_record)
        total_record = (
            pd.Series(first_round_wins_by_record).astype(int).astype(str) + "-" +
            first_round_losses_by_record.astype(int).astype(str)
        )

        # Combine into a DataFrame
        first_round_summary_df = pd.DataFrame({
            "Win Probability": first_round_probabilities,
            "Total Record": total_record,
            "Total Games": first_round_game_counts
        })

        first_round_summary_df["Total Games Played"] = first_round_summary_df.index.map(lambda r: sum(map(int, r.split('-'))))
        first_round_summary_df = first_round_summary_df[first_round_summary_df["Total Games Played"] == 14].copy()
        first_round_summary_df.drop(columns=["Total Games Played"], inplace=True)

        # Ensure the index is sorted by the number of wins in ascending order
        first_round_summary_df.index = pd.Categorical(
            first_round_summary_df.index, 
            categories=sorted(first_round_summary_df.index, key=lambda r: int(r.split('-')[0])),
            ordered=True
        )
        first_round_summary_df = first_round_summary_df.sort_index()
        # Map record indices to "x" format
        first_round_summary_df.index = first_round_summary_df.index.map(lambda x: f"{x}")

        # Print results
        print("First round win probabilities by record:")
        print(first_round_summary_df)
        print()
        # --------------------------------------------------------------------------------------------

        # --------------------------------------------------------------------------------------------
        # Filter second round games
        second_round_games = all_playoff_dfs[all_playoff_dfs['Round'] == 'Semi Final']

        # Initialize counters for the second round wins by record
        second_round_wins_by_record = {}

        # Loop through the second round games and count wins for each record
        for _, game in second_round_games.iterrows():
            # Extract winner and their record (either Record 1 or Record 2)
            winner = game['Winner']
            if winner == game['Team 1']:
                winner_record = game['Record 1']
            else:
                winner_record = game['Record 2']
            
            # Count wins for each record
            if winner_record in second_round_wins_by_record:
                second_round_wins_by_record[winner_record] += 1
            else:
                second_round_wins_by_record[winner_record] = 1

        # Count total appearances of each record in the second round (both Record 1 and Record 2)
        second_round_game_counts = second_round_games['Record 1'].value_counts().add(
            second_round_games['Record 2'].value_counts(), fill_value=0
        ).astype(int)

        # Ensure both indices align, and fill missing values with 0
        all_records = second_round_game_counts.index.union(second_round_wins_by_record.keys())
        second_round_wins_by_record = {record: second_round_wins_by_record.get(record, 0) for record in all_records}
        second_round_game_counts = second_round_game_counts.reindex(all_records, fill_value=0)

        # Calculate second round win probabilities
        second_round_probabilities = (
            pd.Series(second_round_wins_by_record) / second_round_game_counts * 100
        ).fillna(0).round(2).map(lambda x: f"{x:.2f}%")

        # Calculate total record (wins-losses) for each record
        second_round_losses_by_record = second_round_game_counts - pd.Series(second_round_wins_by_record)
        total_record = (
            pd.Series(second_round_wins_by_record).astype(int).astype(str) + "-" +
            second_round_losses_by_record.astype(int).astype(str)
        )

        # Combine into a DataFrame
        second_round_summary_df = pd.DataFrame({
            "Win Probability": second_round_probabilities,
            "Total Record": total_record,
            "Total Games": second_round_game_counts
        })

        # Add and filter by "Total Games Played" (only 14-game records)
        second_round_summary_df["Total Games Played"] = second_round_summary_df.index.map(lambda r: sum(map(int, r.split('-'))))
        second_round_summary_df = second_round_summary_df[second_round_summary_df["Total Games Played"] == 14].copy()
        second_round_summary_df.drop(columns=["Total Games Played"], inplace=True)

        # Ensure the index is sorted by the number of wins in ascending order
        second_round_summary_df.index = pd.Categorical(
            second_round_summary_df.index, 
            categories=sorted(second_round_summary_df.index, key=lambda r: int(r.split('-')[0])),
            ordered=True
        )
        second_round_summary_df = second_round_summary_df.sort_index()

        # Map record indices to "x" format
        second_round_summary_df.index = second_round_summary_df.index.map(lambda x: f"{x}")

        # Print results
        print("Second round win probabilities by record:")
        print(second_round_summary_df)
        print()
        # --------------------------------------------------------------------------------------------

        # --------------------------------------------------------------------------------------------
        # Filter championship games
        championship_games = all_playoff_dfs[all_playoff_dfs['Round'] == 'Championship']

        # Initialize counters for the championship wins by record
        championship_wins_by_record = {}

        # Loop through the championship games and count wins for each record
        for _, game in championship_games.iterrows():
            # Extract winner and their record (either Record 1 or Record 2)
            winner = game['Winner']
            if winner == game['Team 1']:
                winner_record = game['Record 1']
            else:
                winner_record = game['Record 2']
            
            # Count wins for each record
            if winner_record in championship_wins_by_record:
                championship_wins_by_record[winner_record] += 1
            else:
                championship_wins_by_record[winner_record] = 1

        # Count total appearances of each record in the championship round (both Record 1 and Record 2)
        championship_game_counts = championship_games['Record 1'].value_counts().add(
            championship_games['Record 2'].value_counts(), fill_value=0
        ).astype(int)

        # Ensure both indices align, and fill missing values with 0
        all_records = championship_game_counts.index.union(championship_wins_by_record.keys())
        championship_wins_by_record = {record: championship_wins_by_record.get(record, 0) for record in all_records}
        championship_game_counts = championship_game_counts.reindex(all_records, fill_value=0)

        # Calculate championship win probabilities
        championship_probabilities = (
            pd.Series(championship_wins_by_record) / championship_game_counts * 100
        ).fillna(0).round(2).map(lambda x: f"{x:.2f}%")

        # Calculate total record (wins-losses) for each record
        championship_losses_by_record = championship_game_counts - pd.Series(championship_wins_by_record)
        total_record = (
            pd.Series(championship_wins_by_record).astype(int).astype(str) + "-" +
            championship_losses_by_record.astype(int).astype(str)
        )

        # Combine into a DataFrame
        championship_summary_df = pd.DataFrame({
            "Win Probability": championship_probabilities,
            "Total Record": total_record,
            "Total Games": championship_game_counts
        })

        # Add and filter by "Total Games Played" (only 14-game records)
        championship_summary_df["Total Games Played"] = championship_summary_df.index.map(lambda r: sum(map(int, r.split('-'))))
        championship_summary_df = championship_summary_df[championship_summary_df["Total Games Played"] == 14].copy()
        championship_summary_df.drop(columns=["Total Games Played"], inplace=True)

        # Ensure the index is sorted by the number of wins in ascending order
        championship_summary_df.index = pd.Categorical(
            championship_summary_df.index, 
            categories=sorted(championship_summary_df.index, key=lambda r: int(r.split('-')[0])),
            ordered=True
        )
        championship_summary_df = championship_summary_df.sort_index()

        # Map record indices to "x" format
        championship_summary_df.index = championship_summary_df.index.map(lambda x: f"{x}")

        # Print results
        print("Out of " + str(total_seasons)+ " seasons analyzed")
        print("Championship win probabilities by record:")
        print(championship_summary_df)
        print()
        # --------------------------------------------------------------------------------------------

    wins_by_record(all_playoff_dfs)

app()